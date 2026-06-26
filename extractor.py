"""
Extracts structured fleet energy data from any Excel format using Claude Opus 4.8.
The LLM handles the variable schema — we just give it the raw content and ask for JSON.
"""

import openpyxl
import json
import os
import re
import requests

OPENROUTER_KEY = os.environ["OPENROUTER_KEY"]
MODEL = "anthropic/claude-opus-4.8"


def _extract_quarterly_summary(ws) -> dict | None:
    """For a quarterly sheet extract the pre-calculated total row.

    Quarterly sheets have an embedded total row where the plate column (col 3) is
    empty and col 4 holds the fleet-wide km total. This is more reliable than
    summing vehicle rows, which double-counts category subtotals.

    Returns None if no valid total row is found (e.g. all km values are negative
    due to odometer resets — treat as corrupted/unavailable quarter).
    """
    rows = list(ws.iter_rows(values_only=True))

    # Find column indices from header rows
    km_col = fuel_col = cargo_col = tep_col = None
    for row in rows[:5]:
        lower = [str(c).lower().strip() if c else "" for c in row[:25]]
        for j, v in enumerate(lower):
            if v in ("km", "kms") and km_col is None:
                km_col = j
            if "consumo" in v and fuel_col is None:
                fuel_col = j
            if v in ("toneladas", "ton", "toneladas ") and cargo_col is None:
                cargo_col = j
            if v == "tep" and tep_col is None:
                tep_col = j

    if km_col is None:
        return None

    # Look for the total row: col 3 (plate) is None, col km_col is a plausible fleet total
    for row in rows:
        if row[3] is not None:
            continue
        km_val = row[km_col] if km_col < len(row) else None
        if not isinstance(km_val, (int, float)):
            continue
        if km_val < 10_000 or km_val > 10_000_000:
            continue
        fuel_val = row[fuel_col] if fuel_col and fuel_col < len(row) else 0.0
        cargo_val = row[cargo_col] if cargo_col and cargo_col < len(row) else 0.0
        tep_val = row[tep_col] if tep_col and tep_col < len(row) else 0.0
        return {
            "total_km": round(km_val),
            "total_fuel_l": round(float(fuel_val or 0), 2),
            "total_cargo_t": round(float(cargo_val or 0), 2),
            "total_tep": round(float(tep_val or 0), 4),
        }

    return None  # No valid total row — quarter has corrupted data


def excel_to_text(path: str, max_rows_per_sheet: int = 300) -> str:
    """Convert Excel workbook to a compact text representation.

    For quarterly sheets (names containing 'trim'), extracts a clean per-quarter
    summary (km, fuel, cargo) instead of dumping all vehicle rows — this keeps the
    text within budget and avoids corrupted odometer totals skewing the sum.
    Skips sheets with >500 rows (raw operational dumps).
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    parts = []
    quarterly_summaries = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Large raw-data sheets: skip entirely
        if ws.max_row > 500:
            parts.append(f"\n=== SHEET: {sheet_name} (SKIPPED — {ws.max_row} rows, raw data) ===")
            continue

        # Quarterly sheets: extract clean numeric summary, don't dump all rows
        if "trim" in sheet_name.lower() or "trimestre" in sheet_name.lower():
            summary = _extract_quarterly_summary(ws)
            if summary:
                quarterly_summaries.append(f"  {sheet_name}: km={summary['total_km']:,}  fuel_l={summary['total_fuel_l']:,}  cargo_t={summary['total_cargo_t']:,}  tep={summary['total_tep']}")
                continue
            # fallback: include raw rows if extraction failed

        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows_per_sheet:
                rows.append(f"... ({ws.max_row - max_rows_per_sheet} more rows truncated)")
                break
            if any(c is not None for c in row):
                cells = [str(c) if c is not None else "" for c in row[:25]]
                rows.append(" | ".join(cells))
        if rows:
            parts.append(f"\n=== SHEET: {sheet_name} ===\n" + "\n".join(rows))

    if quarterly_summaries:
        header = "\n=== QUARTERLY FLEET SUMMARIES (pre-extracted, odometer errors filtered) ===\n"
        header += "\n".join(quarterly_summaries)
        parts.insert(0, header)

    return "\n".join(parts)[:80000]


EXTRACTION_PROMPT = """You are an expert at reading Portuguese fleet energy audit Excel files.

The data comes from a vehicle fleet management system. Formats vary widely between clients —
different column names, different sheet structures, monthly vs quarterly aggregations, etc.

Your job: extract the key annual energy metrics needed to produce a PRCE
(Plano de Racionalização do Consumo de Energia) report.

Analyze the Excel content below and return a JSON object with this exact structure:

{{
  "annual_summary": {{
    "total_km": <total km driven all vehicles full year, integer>,
    "total_energy_tep": <total energy consumption in tep (tonnes equivalent petroleum), float>,
    "total_cargo_tonnes": <total cargo transported in tonnes, float or null if not applicable>,
    "total_work_tkm": <sum of (km_i × cargo_i) computed per vehicle/category row — do NOT multiply total_km × total_cargo_t. If per-row data is unavailable, set null.>
  }},
  "quarterly": [
    {{
      "label": "Q1",
      "km": <km, integer>,
      "energy_tep": <tep, float>,
      "cargo_t": <tonnes, float or null>,
      "fuel_breakdown": {{
        "diesel_l": <litres, float>,
        "gasoline_l": <litres, float>,
        "gnc_kg": <kg of natural gas, float>,
        "kwh": <electric kWh, float>
      }}
    }}
  ],
  "fleet_composition": [
    {{
      "category": "<category name>",
      "count": <integer>,
      "avg_age_years": <float or null>,
      "period_km": <total km for this category in the data period, integer or null>,
      "period_cargo_t": <total cargo in tonnes for this category in the data period, float or null>,
      "period_energy_tep": <total energy in tep for this category in the data period, float or null>
    }}
  ],
  "fuel_types_used": ["diesel", "gasoline", ...],
  "data_period": "<e.g. 2024 full year, Q1-Q2 2025>",
  "data_quality_notes": "<any issues, inconsistencies, or assumptions made>",
  "fleet_type": "<heavy_cargo | urban_services | mixed_logistics | light_only>"
}}

IMPORTANT CONVERSION FACTORS to use if you need to convert:
- Diesel: 0.873 kgep/L → divide by 1000 to get tep/L
- Gasoline: 0.773 kgep/L
- GNL/GNC: 1.149 tep/tonne
- Electric: 0.290 tep/MWh = 0.000290 tep/kWh

IMPORTANT — work_tkm calculation:
For fleet_composition, populate period_km and period_cargo_t per category so the system can
compute work = sum(category_km × category_cargo_t). This is more accurate than the global product.
GNC is always in kg in Portuguese fleet data, never litres.

IMPORTANT — multi-sheet quarterly files:
Many Portuguese fleet files have separate sheets per quarter named "1º trim", "2º trim", "3º trim", "4º trim"
(or similar). If you see multiple quarterly sheets, SUM their km, energy and cargo across ALL quarters present
to produce the annual_summary totals. Set data_period to reflect what's available, e.g. "2025 (Q1-Q3)"
if three quarters are present, or "2025 full year" if all four are present.

IMPORTANT — partial year data:
If data covers less than a full year (e.g. only Q1, or Q1+Q2), do NOT extrapolate.
Report the actual numbers from the data as-is.
Set data_period to the exact period covered (e.g. "Q1 2025", "Q1-Q3 2025").
Set data_quality_notes to something like "DADOS PARCIAIS: apenas Q1 2025. Para o PRCE definitivo é necessário o ano completo."
Only flag as partial if fewer than 4 quarters are present.
The downstream system will handle the partial-data warning.

If cargo data is unavailable, set cargo fields to null.
Return ONLY the JSON object, no explanation.

EXCEL CONTENT:
{excel_content}
"""  # noqa: placeholder replaced via .replace(), not .format()


def _parse_json_from_llm(content: str) -> dict:
    """Robustly extract JSON from LLM output that may have explanation text or code fences."""
    # Strip code fences first, then brace-walk the result
    content = re.sub(r"```(?:json)?\s*", "", content).replace("```", "").strip()

    start = content.find("{")
    if start != -1:
        depth = 0
        for i, ch in enumerate(content[start:], start):
            if ch == "{":
                depth += 1
            elif ch == "}":
                depth -= 1
                if depth == 0:
                    return json.loads(content[start : i + 1])

    raise ValueError(f"Could not extract JSON from LLM response:\n{content[:500]}")


def extract_fleet_data(excel_path: str, max_rows: int = 300) -> dict:
    """Main entry point: read Excel, call Claude, return structured dict."""
    print(f"  Reading Excel: {excel_path}")
    raw = excel_to_text(excel_path, max_rows_per_sheet=max_rows)
    print(f"  Excel converted to text ({len(raw)} chars), calling Claude Opus 4.8...")

    prompt = EXTRACTION_PROMPT.replace("{excel_content}", raw)

    resp = requests.post(
        "https://openrouter.ai/api/v1/chat/completions",
        headers={
            "Authorization": f"Bearer {OPENROUTER_KEY}",
            "Content-Type": "application/json",
            "HTTP-Referer": "https://wearefractional.ai",
            "X-Title": "AE-PRCE Automation",
        },
        json={
            "model": MODEL,
            "messages": [{"role": "user", "content": prompt}],
            "temperature": 0,
            "max_tokens": 5000,
        },
        timeout=120,
    )
    resp.raise_for_status()

    content = resp.json()["choices"][0]["message"]["content"]
    data = _parse_json_from_llm(content)

    print(f"  Extraction complete. Period: {data.get('data_period', 'unknown')}")
    print(f"  Fleet type: {data.get('fleet_type', 'unknown')}")
    if data.get("data_quality_notes"):
        print(f"  Notes: {data['data_quality_notes'][:200]}")
    return data
